from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = ROOT / "data" / "exports"
OUTPUT_PATH = ROOT / "excel" / "MerchantDashboard.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="173F5F")
CARD_FILL = PatternFill("solid", fgColor="F3F7FB")
SECTION_FILL = PatternFill("solid", fgColor="EAF4EA")
ACCENT_FILL = PatternFill("solid", fgColor="FFF3D6")
ALERT_FILL = PatternFill("solid", fgColor="FDE7E7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=18, bold=True)
SUBTITLE_FONT = Font(size=11, italic=True, color="5F6B7A")
LABEL_FONT = Font(size=10, color="4F5D75", bold=True)
VALUE_FONT = Font(size=16, bold=True, color="173F5F")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7E2EE"),
    right=Side(style="thin", color="D7E2EE"),
    top=Side(style="thin", color="D7E2EE"),
    bottom=Side(style="thin", color="D7E2EE"),
)


def require_exports() -> None:
    expected = [
        "kpi_summary.csv",
        "monthly_revenue.csv",
        "payment_mix.csv",
        "top_merchants.csv",
        "revenue_by_state.csv",
        "delivery_summary.csv",
        "review_distribution.csv",
        "category_performance.csv",
        "merchant_health.csv",
        "intervention_candidates.csv",
        "retention.csv",
    ]
    missing = [name for name in expected if not (EXPORT_DIR / name).exists()]
    if missing:
        missing_text = "\n".join(f"- {name}" for name in missing)
        raise FileNotFoundError(
            "Missing export files. Run python/analysis.py first.\n"
            f"Missing:\n{missing_text}"
        )


def load_exports() -> dict[str, pd.DataFrame]:
    frames = {}
    for csv_path in EXPORT_DIR.glob("*.csv"):
        frames[csv_path.stem] = pd.read_csv(csv_path)
    return frames


def write_dataframe(ws, frame: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for col_offset, column in enumerate(frame.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_offset, value=column)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    for row_offset, row in enumerate(frame.itertuples(index=False), start=start_row + 1):
        for col_offset, value in enumerate(row, start=start_col):
            cell = ws.cell(row=row_offset, column=col_offset, value=value)
            cell.border = THIN_BORDER
    auto_fit(ws)


def auto_fit(ws) -> None:
    for column_cells in ws.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        width = min(max(len(value) for value in values) + 2, 28)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def format_value(metric: str, value) -> str:
    if pd.isna(value):
        return "-"
    value = float(value)
    if "Revenue" in metric or "Value" in metric:
        return f"R$ {value:,.2f}"
    if "Rate" in metric or "Score" in metric:
        return f"{value:,.2f}%" if "Rate" in metric else f"{value:,.2f}"
    if "Days" in metric:
        return f"{value:,.2f} days"
    return f"{value:,.0f}"


def add_title_block(ws) -> None:
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    ws["A1"] = "Merchant Performance Command Center"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A2"] = "Revenue, retention, payments, and delivery health across the Olist marketplace"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left")


def add_kpi_cards(ws, kpi_frame: pd.DataFrame) -> None:
    add_title_block(ws)
    card_metrics = [
        "Total Revenue",
        "Total Orders",
        "Average Order Value",
        "Repeat Customer Rate %",
        "Late Delivery Rate %",
        "Average Review Score",
    ]
    selected = kpi_frame[kpi_frame["metric"].isin(card_metrics)].reset_index(drop=True)
    starts = ["A4", "C4", "E4", "G4", "I4", "A8"]
    for (metric, value), anchor in zip(selected.itertuples(index=False), starts):
        col = anchor[0]
        row = int(anchor[1:])
        end_col = get_column_letter(ws[f"{col}{row}"].column + 1)
        ws.merge_cells(f"{col}{row}:{end_col}{row}")
        ws.merge_cells(f"{col}{row+1}:{end_col}{row+2}")
        for r in range(row, row + 3):
            for c in range(ws[f"{col}{row}"].column, ws[f"{end_col}{row}"].column + 1):
                cell = ws.cell(r, c)
                cell.fill = CARD_FILL
                cell.border = THIN_BORDER
        ws[f"{col}{row}"] = metric
        ws[f"{col}{row}"].font = LABEL_FONT
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        ws[f"{col}{row+1}"] = format_value(metric, value)
        ws[f"{col}{row+1}"].font = VALUE_FONT
        ws[f"{col}{row+1}"].alignment = Alignment(horizontal="center", vertical="center")

    for column in "ABCDEFGHIJ":
        ws.column_dimensions[column].width = 14


def add_monthly_charts(ws, frame: pd.DataFrame) -> None:
    write_dataframe(ws, frame)
    line_chart = LineChart()
    line_chart.title = "Monthly Revenue"
    line_chart.y_axis.title = "Revenue"
    line_chart.x_axis.title = "Month"
    line_data = Reference(ws, min_col=2, min_row=1, max_row=len(frame) + 1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=len(frame) + 1)
    line_chart.add_data(line_data, titles_from_data=True)
    line_chart.set_categories(categories)
    line_chart.height = 8
    line_chart.width = 16
    line_chart.style = 10
    ws.add_chart(line_chart, "F2")


def add_payment_chart(ws, frame: pd.DataFrame) -> None:
    write_dataframe(ws, frame)
    pie = PieChart()
    pie.title = "Payment Method Distribution"
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(frame) + 1)
    data = Reference(ws, min_col=3, min_row=1, max_row=len(frame) + 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12
    pie.style = 10
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "E2")


def add_bar_chart(ws, frame: pd.DataFrame, title: str, category_col: int, value_col: int, anchor: str) -> None:
    write_dataframe(ws, frame)
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Category"
    chart.x_axis.title = "Value"
    data = Reference(ws, min_col=value_col, min_row=1, max_row=len(frame) + 1)
    categories = Reference(ws, min_col=category_col, min_row=2, max_row=len(frame) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 16
    chart.style = 10
    ws.add_chart(chart, anchor)


def add_retention_heatmap(ws, frame: pd.DataFrame) -> None:
    pivot = frame.pivot(index="cohort_month", columns="activity_month", values="retention_rate_pct").fillna(0)
    pivot = pivot.reset_index()
    write_dataframe(ws, pivot)
    max_row = len(pivot) + 1
    max_col = len(pivot.columns)
    ws.conditional_formatting.add(
        f"B2:{get_column_letter(max_col)}{max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="FCE4D6",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="C6E0B4",
        ),
    )


def add_dashboard_tables(ws, frames: dict[str, pd.DataFrame]) -> None:
    ws["A12"] = "Executive Takeaways"
    ws["A12"].font = Font(size=12, bold=True, color="173F5F")
    insights = build_insights(frames)
    fills = [SECTION_FILL, ACCENT_FILL, ALERT_FILL]
    row = 13
    for idx, insight in enumerate(insights):
        ws.merge_cells(f"A{row}:E{row+1}")
        for r in range(row, row + 2):
            for c in range(1, 6):
                cell = ws.cell(r, c)
                cell.fill = fills[idx]
                cell.border = THIN_BORDER
        ws[f"A{row}"] = insight
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        row += 3

    ws["G12"] = "Merchants To Watch"
    ws["G12"].font = Font(size=12, bold=True, color="173F5F")
    intervention = frames["intervention_candidates"].head(5).copy()
    if not intervention.empty:
        intervention = intervention[
            ["seller_id", "revenue", "avg_review_score", "late_delivery_rate_pct"]
        ]
        write_dataframe(ws, intervention, start_row=13, start_col=7)

    ws["G21"] = "Merchant Health Segments"
    ws["G21"].font = Font(size=12, bold=True, color="173F5F")
    segment_summary = (
        frames["merchant_health"]
        .groupby("merchant_segment", as_index=False)
        .agg(
            merchants=("seller_id", "count"),
            avg_health_score=("merchant_health_score", "mean"),
            total_revenue=("revenue", "sum"),
        )
        .sort_values("total_revenue", ascending=False)
    )
    write_dataframe(ws, segment_summary, start_row=22, start_col=7)


def add_dashboard_charts(ws, monthly_ws, payment_ws, merchant_ws) -> None:
    revenue_chart = LineChart()
    revenue_chart.title = "Monthly Revenue Trend"
    revenue_chart.y_axis.title = "Revenue"
    revenue_chart.x_axis.title = "Month"
    revenue_data = Reference(monthly_ws, min_col=2, min_row=1, max_row=monthly_ws.max_row)
    revenue_cats = Reference(monthly_ws, min_col=1, min_row=2, max_row=monthly_ws.max_row)
    revenue_chart.add_data(revenue_data, titles_from_data=True)
    revenue_chart.set_categories(revenue_cats)
    revenue_chart.height = 8
    revenue_chart.width = 15
    revenue_chart.style = 10
    ws.add_chart(revenue_chart, "A24")

    payment_chart = PieChart()
    payment_chart.title = "Payment Mix"
    payment_labels = Reference(payment_ws, min_col=1, min_row=2, max_row=payment_ws.max_row)
    payment_data = Reference(payment_ws, min_col=3, min_row=1, max_row=payment_ws.max_row)
    payment_chart.add_data(payment_data, titles_from_data=True)
    payment_chart.set_categories(payment_labels)
    payment_chart.height = 8
    payment_chart.width = 10
    payment_chart.style = 10
    payment_chart.dataLabels = DataLabelList()
    payment_chart.dataLabels.showPercent = True
    ws.add_chart(payment_chart, "G28")

    merchant_chart = BarChart()
    merchant_chart.title = "Top Merchant Revenue"
    merchant_data = Reference(merchant_ws, min_col=4, min_row=1, max_row=min(merchant_ws.max_row, 11))
    merchant_cats = Reference(merchant_ws, min_col=1, min_row=2, max_row=min(merchant_ws.max_row, 11))
    merchant_chart.add_data(merchant_data, titles_from_data=True)
    merchant_chart.set_categories(merchant_cats)
    merchant_chart.height = 8
    merchant_chart.width = 13
    merchant_chart.style = 10
    ws.add_chart(merchant_chart, "A40")


def build_insights(frames: dict[str, pd.DataFrame]) -> list[str]:
    kpis = dict(frames["kpi_summary"].itertuples(index=False, name=None))
    payment_top = frames["payment_mix"].iloc[0]
    state_top = frames["revenue_by_state"].iloc[0]
    health = frames["merchant_health"]
    at_risk = int((health["merchant_segment"] == "High Value / At Risk").sum())
    return [
        (
            f"Revenue reached R$ {kpis['Total Revenue']:,.0f} across {kpis['Total Orders']:,.0f} orders. "
            f"{payment_top['payment_type'].title()} is the top payment method by revenue."
        ),
        (
            f"{state_top['customer_state']} leads revenue generation, while repeat customers currently account for "
            f"{kpis['Repeat Customer Rate %']:.2f}% of the customer base, leaving retention upside."
        ),
        (
            f"Late deliveries are at {kpis['Late Delivery Rate %']:.2f}% and there are {at_risk} high-value merchants "
            f"that need proactive intervention based on the Merchant Health Score."
        ),
    ]


def add_notes_sheet(ws) -> None:
    ws["A1"] = "How to finish the Excel deliverable"
    ws["A1"].font = TITLE_FONT
    notes = [
        "Refresh the workbook after rerunning the Python analysis exports.",
        "Convert data sheets to Excel Tables for recruiter-friendly filtering.",
        "If using desktop Excel, build Pivot Tables from the data sheets and add slicers for month, state, and payment type.",
        "Create a timeline control on purchase month for a more business-ready demo.",
        "Link this workbook to screenshots in the README and to the Power BI file if you build one.",
        "Use the Merchant Health sheet to build account-prioritization talking points for DotPe-style interviews.",
    ]
    for idx, note in enumerate(notes, start=3):
        ws[f"A{idx}"] = note
        ws[f"A{idx}"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 100


def main() -> None:
    require_exports()
    frames = load_exports()
    workbook = Workbook()

    dashboard = workbook.active
    dashboard.title = "Dashboard"
    add_kpi_cards(dashboard, frames["kpi_summary"])

    monthly = workbook.create_sheet("MonthlyRevenue")
    add_monthly_charts(monthly, frames["monthly_revenue"])

    payments = workbook.create_sheet("PaymentMix")
    add_payment_chart(payments, frames["payment_mix"])

    merchants = workbook.create_sheet("TopMerchants")
    add_bar_chart(merchants, frames["top_merchants"], "Top Merchants by Revenue", 1, 4, "G2")

    states = workbook.create_sheet("RevenueByState")
    add_bar_chart(states, frames["revenue_by_state"], "Revenue by State", 1, 2, "F2")

    delivery = workbook.create_sheet("DeliveryOps")
    add_bar_chart(delivery, frames["delivery_summary"], "Late Delivery Rate % by Month", 1, 4, "F2")

    reviews = workbook.create_sheet("Reviews")
    add_bar_chart(reviews, frames["review_distribution"], "Review Score Distribution", 1, 2, "E2")

    categories = workbook.create_sheet("Categories")
    add_bar_chart(categories, frames["category_performance"], "Top Categories by Revenue", 1, 3, "F2")

    health = workbook.create_sheet("MerchantHealth")
    write_dataframe(health, frames["merchant_health"])
    health.conditional_formatting.add(
        f"G2:G{health.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="FDE7E7",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF3D6",
            end_type="max",
            end_color="D9F0D8",
        ),
    )

    intervention = workbook.create_sheet("AtRiskMerchants")
    write_dataframe(intervention, frames["intervention_candidates"])

    retention = workbook.create_sheet("Retention")
    add_retention_heatmap(retention, frames["retention"])

    notes = workbook.create_sheet("Notes")
    add_notes_sheet(notes)

    add_dashboard_tables(dashboard, frames)
    add_dashboard_charts(dashboard, monthly, payments, merchants)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(f"Excel dashboard workbook created at: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
